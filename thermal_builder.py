import os
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template_files', 'template_thermal.xlsx')

# ব্যবহারকারীর স্পষ্ট নির্দেশ অনুযায়ী — Gmt. Color, Gmt. Size, Qty মাস্ট-হ্যাভ।
# Measurement ইচ্ছাকৃতভাবে এখানে নেই (সেটা আলাদা confirm-ডায়ালগ দিয়ে হ্যান্ডেল হয়) —
# UI-তে ফাঁকা রেখে "ইয়েস" করলে blank measurement নিয়েই প্রসিড করা হবে, এটা
# ভ্যালিডেশন-এরর/warning হিসেবে গণ্য হবে না।
REQUIRED_FIELDS = ['color', 'size', 'qty']


def to_num(v):
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return v


def na_if_blank(v):
    v = str(v).strip() if v is not None else ''
    return v if v else 'N/A'


def validate_thermal_line_items(line_items):
    """মাস্ট-হ্যাভ ফিল্ড (Gmt. Color/Gmt. Size/Qty) মিসিং থাকলে warning রিটার্ন করে।"""
    warnings = []
    for i, item in enumerate(line_items):
        missing = [f for f in REQUIRED_FIELDS if not str(item.get(f, '')).strip()]
        if missing:
            warnings.append(f"Row {i + 1}: {', '.join(missing)} খালি আছে — চেক করুন")
    return warnings


def _strip_external_links(wb):
    """Carton বিল্ডারের মতোই — Excel 'Repairs to...' ওয়ার্নিং এড়াতে যেকোনো
    external link স্থায়ীভাবে মুছে দেওয়া হয়।"""
    try:
        if getattr(wb, '_external_links', None):
            wb._external_links = []
    except Exception:
        pass
    try:
        if hasattr(wb, '_external_references'):
            wb._external_references = []
    except Exception:
        pass


def _write_df_sheet(wb, sheet_name, df):
    if df is None or df.empty:
        return
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_thermal_excel(line_items, header_info, out_path,
                         customer_override=None, buyer_override=None,
                         po_override=None, delivery_date='', delivery_address='',
                         measurement='', raw_df=None, summary_df=None, warnings=None,
                         remark_place=False, remark_address=False):
    """Thermal মডিউলের জন্য একটাই .xlsx ফাইল বানায়। শীটগুলো Carton-এর মতোই:
    - Sheet1     : Mapped Template
    - Raw Data   : canonical (মেল্ট করা, প্রতি সাইজ = এক রো) লাইন-আইটেম
    - PO Details : PDF-এ যেভাবে ছিল (wide ফরম্যাট, অরিজিনাল কলাম)
    - PO Summary : PDF-এর page0-এর ST Caow/Rate সামারি টেবিল
    - Warnings   : মাস্ট-হ্যাভ ফিল্ড মিসিং থাকলে

    measurement: ম্যানুয়ালি ইনপুট দেওয়া মেজারমেন্ট — সব রো-তে এক ভ্যালুই বসবে।
    ফাঁকা রাখা হলে (ইউজার confirm করার পর) Measurement কলাম ফাঁকাই থাকবে,
    এটা warning তৈরি করবে না।
    """
    wb = load_workbook(TEMPLATE_PATH)
    _strip_external_links(wb)
    ws = wb['Sheet1']

    ws['B2'] = po_override or header_info.get('po_number', '') or 'N/A'
    ws['B3'] = customer_override or header_info.get('customer', '') or 'N/A'
    ws['B4'] = buyer_override or header_info.get('buyer', '') or 'N/A'

    sample_row = 8
    n_cols = 22  # A থেকে V পর্যন্ত
    style_ref = [copy(ws.cell(row=sample_row, column=c).font) for c in range(1, n_cols + 1)]
    fill_ref = [copy(ws.cell(row=sample_row, column=c).fill) for c in range(1, n_cols + 1)]
    align_ref = [copy(ws.cell(row=sample_row, column=c).alignment) for c in range(1, n_cols + 1)]
    border_ref = [copy(ws.cell(row=sample_row, column=c).border) for c in range(1, n_cols + 1)]
    numfmt_ref = [ws.cell(row=sample_row, column=c).number_format for c in range(1, n_cols + 1)]

    start_row = 8
    for i, r in enumerate(line_items):
        row = start_row + i

        # Remarks কলামে একটাই কলাম আছে (Carton-এর মতো দুইটা আলাদা কলাম নেই),
        # তাই দুটো চেকবক্স একসাথে চেক করা থাকলে " | " দিয়ে জোড়া লাগিয়ে বসানো হবে।
        remark_parts = []
        if remark_place and r.get('delivery_place_pdf'):
            remark_parts.append(f"Delivery Place: {r.get('delivery_place_pdf')}")
        if remark_address and r.get('delivery_address_pdf'):
            remark_parts.append(f"Delivery Address: {r.get('delivery_address_pdf')}")
        remarks_val = ' | '.join(remark_parts)

        values = [
            'Thermal Sticker',                     # A Item Name (ফিক্সড)
            measurement or '',                       # B Measurement (ম্যানুয়াল ইনপুট)
            na_if_blank(r.get('ewo_no')),              # C Gmt. EWO No
            na_if_blank(r.get('style_no')),              # D Gmt. Style No
            na_if_blank(r.get('po_no')),                   # E Gmt. PO
            'All',                                           # F Gmt. Destination
            na_if_blank(r.get('reference')),                  # G Reference/SKU Number (PDF Instruction কলাম থেকে)
            'N/A',                                              # H Pack Type
            na_if_blank(r.get('color')),                         # I Gmt. Color
            na_if_blank(r.get('size')),                            # J Gmt. Size
            'N/A',                                                    # K Paper GSM
            'N/A',                                                      # L Paper Type
            'N/A',                                                        # M Color
            'N/A',                                                          # N Type of Print
            'N/A',                                                            # O No Of Print Color
            'N/A',                                                              # P Thermal FG No Of Sheet
            r.get('uom') or 'Pcs',                                                # Q UOM
            to_num(r.get('qty', '')),                                              # R Order Qty
            to_num(r.get('rate', '')) if str(r.get('rate', '')).strip() else '',     # S Rate($)
            delivery_date or '',                                                       # T Delivery Date
            delivery_address or '',                                                      # U Delivery Place
            remarks_val,                                                                   # V Remarks
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            style_c = min(c, n_cols) - 1
            cell.font = style_ref[style_c]
            cell.fill = fill_ref[style_c]
            cell.alignment = align_ref[style_c]
            cell.border = border_ref[style_c]
            cell.number_format = numfmt_ref[style_c]

    _write_df_sheet(wb, 'Raw Data', pd.DataFrame(line_items))
    _write_df_sheet(wb, 'PO Details', raw_df)
    _write_df_sheet(wb, 'PO Summary', summary_df)
    if warnings:
        _write_df_sheet(wb, 'Warnings', pd.DataFrame({'Warning': warnings}))

    wb.active = 0
    wb.save(out_path)
