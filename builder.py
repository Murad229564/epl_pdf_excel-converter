import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from config import ITEM_NAME_ALIASES, resolve_alias

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'template_files')
TEMPLATE_PATHS = {
    'IN-HOUSE': os.path.join(TEMPLATE_DIR, 'template_inhouse.xlsx'),
    'OUT-HOUSE': os.path.join(TEMPLATE_DIR, 'template_general.xlsx'),
}

REQUIRED_FIELDS = ['item_name', 'ewo_no', 'style_no', 'length', 'width', 'height', 'ply', 'qty']

# এই আইটেমগুলোতে সাধারণত Height থাকে না (Divider, Top Bottom) —
# তাই এগুলোর ক্ষেত্রে Height মিসিং থাকলে warning দেওয়া হবে না,
# বরং Height-এ ভ্যালু পাওয়া গেলে (যেটা হওয়ার কথা না) warning দেওয়া হবে।
HEIGHT_EXEMPT_KEYWORDS = ['divider', 'top bottom', 'top-bottom', 'top/bottom', 'cover top']


def is_height_exempt(item_name):
    name = str(item_name or '').lower()
    return any(k in name for k in HEIGHT_EXEMPT_KEYWORDS)


def to_num(v):
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return v


def na_if_blank(v):
    v = str(v).strip() if v is not None else ''
    return v if v else 'N/A'


def validate_line_items(line_items):
    """Returns a list of warning strings for rows missing a MUST-HAVE field.
    Divider / Top Bottom জাতীয় আইটেমের ক্ষেত্রে Height মাস্ট-হ্যাভ না —
    উল্টো Height-এ ভ্যালু থাকলে সেটাই ফ্ল্যাগ করা হয়।"""
    warnings = []
    for i, item in enumerate(line_items):
        exempt = is_height_exempt(item.get('item_name'))
        required = [f for f in REQUIRED_FIELDS if f != 'height'] if exempt else REQUIRED_FIELDS
        missing = [f for f in required if not str(item.get(f, '')).strip()]
        if missing:
            warnings.append(f"Row {i + 1}: {', '.join(missing)} খালি আছে — চেক করুন")
        if exempt and str(item.get('height', '')).strip():
            warnings.append(
                f"Row {i + 1}: '{item.get('item_name')}' আইটেমে সাধারণত Height থাকার কথা না, "
                f"কিন্তু একটা ভ্যালু পাওয়া গেছে — চেক করুন"
            )
    return warnings


def _write_df_sheet(wb, sheet_name, df):
    """একটা pandas DataFrame-কে নতুন শীট হিসেবে workbook-এ header-সহ বসিয়ে দেয়।"""
    if df is None or df.empty:
        return
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    # হেডার রো বোল্ড করে দেওয়া
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _strip_external_links(wb):
    """Excel-এ ফাইল ওপেন করলে মাঝেমধ্যে 'We found a problem with some content'
    / 'Repairs to ...xlsx' ওয়ার্নিং আসার প্রধান কারণ হলো ওয়ার্কবুকের ভেতরে
    অন্য কোনো external xlsx ফাইলের রেফারেন্স (externalLinks) থেকে যাওয়া —
    এটা সাধারণত টেমপ্লেট ফাইলটা কখনো অন্য কোনো ওয়ার্কবুক থেকে কপি-পেস্ট বা
    ফর্মুলা-লিংক করে বানানোর সময় ঢুকে যায়। এই ফাংশনটা সবসময় (প্রতিটা আউটপুট
    ফাইল বানানোর সময়) যেকোনো external link/reference মুছে দেয়, টেমপ্লেটে এটা
    থাকুক বা না থাকুক — যাতে ভবিষ্যতে টেমপ্লেট আপডেট হলেও এই সমস্যা আর
    কখনো ফিরে না আসে।"""
    try:
        if getattr(wb, '_external_links', None):
            wb._external_links = []
    except Exception:
        pass
    try:
        if hasattr(wb, '_external_references'):
            wb._external_references = []
    except Exception:
        pass


def _matches_u_divider_measurement(length, width):
    """MARKS & SPENCER SCM LTD.-এর জন্য একটা স্পেশাল বিজনেস রুল: Length/Width
    এই দুটো নির্দিষ্ট কম্বিনেশনের একটার সাথে মিললে Item Name 'U Divider' হবে
    (অন্য কোনো buyer-এর ক্ষেত্রে এটা প্রযোজ্য না)।"""
    try:
        l = round(float(str(length).replace(',', '').strip()))
        w = round(float(str(width).replace(',', '').strip()))
    except (ValueError, TypeError):
        return False
    return (l, w) in {(93, 36), (51, 41)}


def build_combined_excel(line_items, header_info, out_path, profile='IN-HOUSE',
                          customer_override=None, buyer_override=None,
                          po_override=None, delivery_date='', delivery_address='',
                          raw_df=None, summary_df=None, warnings=None,
                          remark_place=False, remark_address=False):
    """একটাই .xlsx ফাইল বানায় — zip আর দরকার নেই। শীটগুলো:
    - Sheet1        : Mapped Template (এটাই ওপেন হলে সবার আগে দেখা যাবে)
    - Raw Data      : canonical raw extracted line items
    - PO Details    : PDF-এ যেভাবে কলাম ছিল ঠিক সেভাবে (অরিজিনাল হেডিং)
    - PO Summary    : PDF-এর প্রথম পাতার Business Line সামারি টেবিল (থাকলে)
    - Warnings      : must-have ফিল্ড মিসিং থাকলে (থাকলে)

    remark_place / remark_address: UI-র দুটো চেকমার্ক। চেক করা থাকলে PDF থেকে
    পাওয়া সেই row-এর Delivery Place / Delivery Address ভ্যালুটা Remarks কলামে
    (এবং তার পরের কলামে) বসে যাবে। দুটো আনচেক থাকলে (ডিফল্ট) Remarks ফাঁকাই থাকবে।
    """
    wb = load_workbook(TEMPLATE_PATHS[profile])
    _strip_external_links(wb)
    ws = wb['Sheet1']

    ws['B2'] = po_override or header_info.get('po_number', '') or 'N/A'
    ws['B3'] = customer_override or header_info.get('customer', '') or 'N/A'
    effective_buyer = buyer_override or header_info.get('buyer', '') or ''
    ws['B4'] = effective_buyer or 'N/A'

    sample_row = 8
    n_cols = 19
    style_ref = [copy(ws.cell(row=sample_row, column=c).font) for c in range(1, n_cols + 1)]
    fill_ref = [copy(ws.cell(row=sample_row, column=c).fill) for c in range(1, n_cols + 1)]
    align_ref = [copy(ws.cell(row=sample_row, column=c).alignment) for c in range(1, n_cols + 1)]
    border_ref = [copy(ws.cell(row=sample_row, column=c).border) for c in range(1, n_cols + 1)]
    numfmt_ref = [ws.cell(row=sample_row, column=c).number_format for c in range(1, n_cols + 1)]

    # In-House প্রোফাইলে Color/Size বসানোর দরকার নেই — PDF-এ থাকলেও সবসময় N/A
    force_na = profile == 'IN-HOUSE'

    # Remarks (S) ও তার পরের কলাম (T) — শুধু চেকমার্ক করা থাকলেই ব্যবহার হবে।
    # T কলামে যদি এখনো কোনো হেডার না থাকে, একটা হেডার বসিয়ে দেওয়া হচ্ছে
    # (টেমপ্লেটের বিদ্যমান কোনো কলাম/ফরম্যাট ভাঙা হচ্ছে না)।
    if remark_address and not ws.cell(row=7, column=20).value:
        ws.cell(row=7, column=20, value='Delivery Address')

    start_row = 8
    for i, r in enumerate(line_items):
        row = start_row + i
        remarks_val = na_if_blank(r.get('delivery_place_pdf')) if remark_place else ''
        address_remark_val = na_if_blank(r.get('delivery_address_pdf')) if remark_address else ''

        item_name_val = na_if_blank(resolve_alias(r.get('item_name'), ITEM_NAME_ALIASES))
        if effective_buyer == 'MARKS & SPENCER SCM LTD.' and _matches_u_divider_measurement(
                r.get('length', ''), r.get('width', '')):
            item_name_val = 'U Divider'

        values = [
            item_name_val,                            # A Item Name (ERP নামে অটো-কারেক্ট + M&S স্পেশাল রুল)
            na_if_blank(r.get('ewo_no')),            # B Gmt. EWO No
            na_if_blank(r.get('style_no')),          # C Gmt. Style No
            na_if_blank(r.get('po_no')),             # D Gmt. PO
            'All',                                    # E Gmt. Destination
            na_if_blank(r.get('reference')),           # F Reference/SKU Number
            na_if_blank(r.get('pack_type')),             # G Pack Type
            'N/A' if force_na else na_if_blank(r.get('color')),   # H Gmt. Color
            'N/A' if force_na else na_if_blank(r.get('size')),    # I Gmt. Size
            r.get('measurement_unit', 'Cm') or 'Cm',           # J Measurement Unit
            to_num(r.get('length', '')),                        # K Length
            to_num(r.get('width', '')),                           # L Width
            to_num(r.get('height', '')),                            # M Height
            to_num(r.get('ply', '')),                                 # N Ply
            'Pcs',                                                       # O Order Unit
            to_num(r.get('qty', '')),                                     # P Order Qty
            delivery_date or '',                                            # Q Delivery Date
            delivery_address or '',                                          # R Delivery Place
            remarks_val,                                                       # S Remarks (checkbox-এ Delivery Place)
        ]
        if remark_address:
            values.append(address_remark_val)  # T Delivery Address (checkbox-এ, PDF থেকে)
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            style_c = min(c, n_cols) - 1
            cell.font = style_ref[style_c]
            cell.fill = fill_ref[style_c]
            cell.alignment = align_ref[style_c]
            cell.border = border_ref[style_c]
            cell.number_format = numfmt_ref[style_c]

    import pandas as pd
    _write_df_sheet(wb, 'Raw Data', pd.DataFrame(line_items))
    _write_df_sheet(wb, 'PO Details', raw_df)
    _write_df_sheet(wb, 'PO Summary', summary_df)
    if warnings:
        _write_df_sheet(wb, 'Warnings', pd.DataFrame({'Warning': warnings}))

    wb.active = 0  # Excel খুললে যেন Sheet1 (Mapped Template) সবার আগে দেখায়
    wb.save(out_path)
